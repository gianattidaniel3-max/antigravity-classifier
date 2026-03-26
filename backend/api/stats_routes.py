"""
Stats, audit trail, and batch document export endpoints.

GET  /api/stats                    — global dashboard numbers
GET  /api/audit                    — paginated verification history
GET  /api/documents/export-csv     — all verified docs as CSV (optional date filter)
GET  /api/documents/export-xlsx    — same as CSV but Excel format
GET  /api/documents/{doc_id}/export-pdf  — single-document summary PDF
"""
import io
import csv
import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from backend.db.session import get_db
from backend.db.models import Document, DocumentStatus, Case, CaseStatus, VerificationLog, User
from backend.auth.deps import get_current_user

router = APIRouter(dependencies=[Depends(get_current_user)])


# ─────────────────────────────────────────────────────────────────
# Stats
# ─────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    """Global dashboard numbers."""

    # --- document counts by status ---
    status_counts = {s.value: 0 for s in DocumentStatus}
    for row in db.query(Document.status, func.count()).group_by(Document.status).all():
        status_counts[row[0].value] = row[1]
    total_docs = sum(status_counts.values())

    # --- distribution by label (verified docs only) ---
    label_rows = (
        db.query(Document.extracted_label, func.count())
        .filter(Document.status == DocumentStatus.VERIFIED, Document.extracted_label != None)
        .group_by(Document.extracted_label)
        .order_by(func.count().desc())
        .limit(10)
        .all()
    )
    by_label = {row[0]: row[1] for row in label_rows}

    # --- accuracy rate: verifications where label was NOT corrected ---
    total_verifications = db.query(VerificationLog).count()
    correct_verifications = db.query(VerificationLog).filter(
        VerificationLog.label_changed == False
    ).count()
    accuracy_rate = (
        round(correct_verifications / total_verifications, 4)
        if total_verifications > 0 else None
    )

    # --- docs uploaded per day — last 30 days ---
    cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=30)
    daily_rows = (
        db.query(
            func.date(Document.upload_date).label("day"),
            func.count().label("count"),
        )
        .filter(Document.upload_date >= cutoff)
        .group_by(func.date(Document.upload_date))
        .order_by(func.date(Document.upload_date))
        .all()
    )
    docs_per_day = [{"date": str(r.day), "count": r.count} for r in daily_rows]

    # --- cases ---
    total_cases = db.query(Case).count()
    open_cases  = db.query(Case).filter(Case.status == CaseStatus.OPEN).count()

    return {
        "total_documents":    total_docs,
        "by_status":          status_counts,
        "by_label":           by_label,
        "accuracy_rate":      accuracy_rate,
        "total_verifications": total_verifications,
        "docs_per_day":       docs_per_day,
        "total_cases":        total_cases,
        "open_cases":         open_cases,
    }


# ─────────────────────────────────────────────────────────────────
# Audit trail
# ─────────────────────────────────────────────────────────────────

@router.get("/audit")
def get_audit(
    limit:  int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """Paginated verification history, newest first."""
    rows = (
        db.query(VerificationLog)
        .order_by(VerificationLog.verified_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    total = db.query(func.count(VerificationLog.id)).scalar()

    return {
        "total":  total,
        "offset": offset,
        "items": [
            {
                "id":             r.id,
                "document_id":    r.document_id,
                "filename":       r.document.filename if r.document else None,
                "user_email":     r.user.email if r.user else "—",
                "verified_at":    r.verified_at.isoformat(),
                "original_label": r.original_label,
                "final_label":    r.final_label,
                "label_changed":  r.label_changed,
                "fields_changed": r.fields_changed or {},
            }
            for r in rows
        ],
    }


# ─────────────────────────────────────────────────────────────────
# Batch document export
# ─────────────────────────────────────────────────────────────────

def _query_docs(db: Session, date_from: Optional[str], date_to: Optional[str]):
    q = db.query(Document).filter(Document.status == DocumentStatus.VERIFIED)
    if date_from:
        try:
            q = q.filter(Document.upload_date >= datetime.datetime.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format, use YYYY-MM-DD")
    if date_to:
        try:
            dt_to = datetime.datetime.fromisoformat(date_to) + datetime.timedelta(days=1)
            q = q.filter(Document.upload_date < dt_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format, use YYYY-MM-DD")
    return q.order_by(Document.upload_date.desc()).all()


_CSV_COLUMNS = [
    "filename", "upload_date", "label", "category", "date",
    "confidence", "importo", "mittente", "destinatario",
    "oggetto", "scadenza", "tribunale", "numero_decreto", "numero_rg",
]


def _doc_row(doc: Document) -> list:
    f = doc.extracted_fields or {}
    return [
        doc.filename,
        doc.upload_date.strftime("%d/%m/%Y %H:%M") if doc.upload_date else "",
        doc.extracted_label or "",
        doc.extracted_category or "",
        doc.extracted_date or "",
        f"{doc.confidence_score * 100:.1f}%" if doc.confidence_score else "",
        f.get("importo", ""),
        f.get("mittente", ""),
        f.get("destinatario", ""),
        f.get("oggetto", ""),
        f.get("scadenza", ""),
        f.get("tribunale", ""),
        f.get("numero_decreto", ""),
        f.get("numero_rg", ""),
    ]


@router.get("/documents/export-csv")
def export_docs_csv(
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Download all verified documents as CSV, with optional date range filter."""
    docs = _query_docs(db, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_CSV_COLUMNS)
    for doc in docs:
        writer.writerow(_doc_row(doc))

    filename = f"documenti_{datetime.date.today().isoformat()}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),  # utf-8-sig for Excel compatibility
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/documents/export-xlsx")
def export_docs_xlsx(
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Download all verified documents as an Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    docs = _query_docs(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Documenti Verificati"

    # Header row
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    labels = [c.replace("_", " ").title() for c in _CSV_COLUMNS]
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, doc in enumerate(docs, 2):
        for col_idx, value in enumerate(_doc_row(doc), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"documenti_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────────────────────────
# Per-document PDF summary sheet
# ─────────────────────────────────────────────────────────────────

@router.get("/documents/{doc_id}/export-pdf")
def export_doc_pdf(doc_id: str, db: Session = Depends(get_db)):
    """Generate a one-page PDF summary sheet for a single document."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

    doc = db.query(Document).filter(Document.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("T", parent=styles["Title"],
                                 fontSize=16, textColor=colors.HexColor("#18181b"), spaceAfter=2)
    sub_style   = ParagraphStyle("S", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#71717a"), spaceAfter=2)
    label_style = ParagraphStyle("L", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#52525b"))
    value_style = ParagraphStyle("V", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#18181b"))

    story = []
    story.append(Paragraph("ECO — Extractor", title_style))
    story.append(Paragraph("Scheda Documento", sub_style))
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor("#e2e8f0"), spaceAfter=10))

    fields_data = doc.extracted_fields or {}
    confidence  = f"{doc.confidence_score * 100:.1f}%" if doc.confidence_score else "—"
    upload_date = doc.upload_date.strftime("%d/%m/%Y %H:%M") if doc.upload_date else "—"
    verified    = "Sì" if doc.human_verified else "No"

    rows = [
        ["File",         doc.filename],
        ["Tipo",         doc.extracted_label or "—"],
        ["Categoria",    doc.extracted_category or "—"],
        ["Data doc.",    doc.extracted_date or "—"],
        ["Confidenza",   confidence],
        ["Caricato il",  upload_date],
        ["Verificato",   verified],
    ]
    for field_key, field_label in [
        ("importo", "Importo"), ("mittente", "Mittente"),
        ("destinatario", "Destinatario"), ("oggetto", "Oggetto"),
        ("scadenza", "Scadenza"), ("tribunale", "Tribunale"),
        ("numero_decreto", "N. Decreto"), ("numero_rg", "N. R.G."),
    ]:
        val = fields_data.get(field_key)
        if val:
            rows.append([field_label, val])

    table_data = [
        [Paragraph(k, label_style), Paragraph(v, value_style)]
        for k, v in rows
    ]
    t = Table(table_data, colWidths=[4*cm, 13*cm])
    t.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS",(0, 0), (-1, -1),
         [colors.HexColor("#f8fafc"), colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Generato il {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} — ECO Extractor",
        sub_style,
    ))

    pdf.build(story)
    buf.seek(0)

    safe_name = doc.filename.replace(".pdf", "") or doc_id[:8]
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=scheda_{safe_name}.pdf"},
    )
